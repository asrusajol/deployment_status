"""Renders DeploymentStatusRow lists (app/services/dashboard.py) to an .xlsx workbook, for
the "Export to Excel" button on both dashboards."""

from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.client_version_status import ClientVersionStatus
from app.services.dashboard import DeploymentStatusRow

COLUMNS = [
    ("Client", lambda r: r.client_name),
    ("System", lambda r: r.environment.capitalize() if r.environment else ""),
    ("URL", lambda r: r.server or ""),
    ("Task ID", lambda r: r.task_id or ""),
    ("Branch", lambda r: r.git_branch or ""),
    ("Commit", lambda r: r.commit_hash or ""),
    ("Version", lambda r: r.version or ""),
    ("Changes", lambda r: r.changes_description or ""),
    ("Requested By", lambda r: r.requested_by or ""),
    ("Approved By", lambda r: r.approved_by or ""),
    ("Deployed By", lambda r: r.deployed_by or ""),
    ("Requested At", lambda r: r.requested_at.strftime("%Y-%m-%d %H:%M UTC") if r.requested_at else ""),
    ("Deployed At", lambda r: r.deployed_at.strftime("%Y-%m-%d %H:%M UTC") if r.deployed_at else ""),
]


def _columns_to_xlsx(rows: list, columns: list[tuple[str, Callable]], sheet_title: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]  # Excel's hard limit on sheet-name length

    headers = [label for label, _ in columns]
    sheet.append(headers)
    for row in rows:
        sheet.append([getter(row) for _, getter in columns])

    for index, (header, getter) in enumerate(columns, start=1):
        widest = max([len(header)] + [len(str(getter(row))) for row in rows])
        sheet.column_dimensions[get_column_letter(index)].width = min(widest + 2, 40)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def rows_to_xlsx(rows: list[DeploymentStatusRow], sheet_title: str) -> bytes:
    return _columns_to_xlsx(rows, COLUMNS, sheet_title)


RELEASE_TRACKER_COLUMNS = [
    ("Client", lambda r: r.client.name if r.client else ""),
    ("Test Current Version", lambda r: r.test_current_version or ""),
    ("Test Updated At", lambda r: r.test_updated_at.strftime("%Y-%m-%d %H:%M UTC") if r.test_updated_at else ""),
    ("Live Current Version", lambda r: r.live_current_version or ""),
    ("Live Updated At", lambda r: r.live_updated_at.strftime("%Y-%m-%d %H:%M UTC") if r.live_updated_at else ""),
    (
        "Main Version",
        lambda r: f"{r.main_version} (PR #{r.main_pr_number})" if r.main_version and r.main_pr_number
        else (r.main_version or ""),
    ),
    ("Main Updated At", lambda r: r.main_updated_at.strftime("%Y-%m-%d %H:%M UTC") if r.main_updated_at else ""),
]


def release_tracker_rows_to_xlsx(rows: list[ClientVersionStatus], sheet_title: str) -> bytes:
    return _columns_to_xlsx(rows, RELEASE_TRACKER_COLUMNS, sheet_title)
