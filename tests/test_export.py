from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.models.bitbucket_main_branch_status import BitbucketMainBranchStatus
from app.models.client import Client
from app.models.client_version_status import ClientVersionStatus
from app.services.export import release_tracker_rows_to_xlsx


def test_release_tracker_rows_to_xlsx_writes_expected_columns():
    row = ClientVersionStatus(
        id=1, client_id=1,
        test_current_version="1.0", test_updated_at=datetime(2026, 8, 27, 9, 0, tzinfo=timezone.utc),
        live_current_version="2026.34.34", live_updated_at=datetime(2026, 8, 27, 10, 0, tzinfo=timezone.utc),
    )
    row.client = Client(name="CRM")
    # Main Version is now the live bitbucket_main_branch_status cache, shared across
    # every row — not a per-row snapshot — so it's passed in separately.
    main_status = BitbucketMainBranchStatus(
        id=1, version="2026.34.40", pr_number=15009,
        version_changed_at=datetime(2026, 8, 20, tzinfo=timezone.utc),
    )

    content = release_tracker_rows_to_xlsx([row], "Release Tracker", main_status=main_status)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    header_row = [cell.value for cell in sheet[1]]
    assert header_row == [
        "Client", "Test Current Version", "Test Updated At",
        "Live Current Version", "Live Updated At",
        "Main Version", "Main Updated At",
    ]
    data_row = [cell.value for cell in sheet[2]]
    assert data_row == [
        "CRM", "1.0", "2026-08-27 09:00 UTC",
        "2026.34.34", "2026-08-27 10:00 UTC",
        "2026.34.40 (PR #15009)", "2026-08-20 00:00 UTC",
    ]


def test_release_tracker_rows_to_xlsx_two_clients_share_the_same_main_version():
    row1 = ClientVersionStatus(id=1, client_id=1)
    row1.client = Client(name="Acme")
    row2 = ClientVersionStatus(id=2, client_id=2)
    row2.client = Client(name="Zebra Corp")
    main_status = BitbucketMainBranchStatus(id=1, version="2026.34.40", pr_number=15009)

    content = release_tracker_rows_to_xlsx([row1, row2], "Release Tracker", main_status=main_status)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    main_version_col = [cell.value for cell in sheet[1]].index("Main Version") + 1
    assert sheet.cell(row=2, column=main_version_col).value == "2026.34.40 (PR #15009)"
    assert sheet.cell(row=3, column=main_version_col).value == "2026.34.40 (PR #15009)"


def test_release_tracker_rows_to_xlsx_without_a_bitbucket_sync_yet():
    row = ClientVersionStatus(id=1, client_id=1)
    row.client = Client(name="CRM")

    content = release_tracker_rows_to_xlsx([row], "Release Tracker", main_status=None)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    data_row = [cell.value for cell in sheet[2]]
    # openpyxl round-trips an empty-string cell write as None on read — this just
    # confirms it didn't error and didn't write a stray "None" string into the cell.
    assert data_row[-2:] == [None, None]  # Main Version, Main Updated At
