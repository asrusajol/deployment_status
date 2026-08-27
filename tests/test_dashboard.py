import json
import re
from datetime import datetime, timedelta, timezone

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401 — registers all models on Base.metadata
from app.config import get_settings
from app.database import Base
from app.models.bitbucket_main_branch_status import BitbucketMainBranchStatus
from app.models.client import Client
from app.models.client_version_record import ClientVersionRecord
from app.models.deployable_task import DeployableTask
from app.models.deployment_execution import DeploymentExecution, ExecutionStatus
from app.models.deployment_request import DeploymentEnvironment, DeploymentRequest, RequestStatus, RequestType
from app.models.user import User, UserRole
from app.services.dashboard import current_deployment_status, deployment_history
from tests.conftest import DEFAULT_TEST_PASSWORD, login_as, make_user

# Matches whatever machine_group_id the app is actually configured to treat as "the
# deploy team" (task_api_deployable_machine_group_id, default 13 / "Team Rajib") — this
# is who's allowed to deploy (require_deploy_team_member), a *different* axis from who's
# allowed to approve (can_approve_deployment_request, which follows the REQUESTER's own
# team instead — see the real bug this was fixed for: a team lead of an unrelated team
# couldn't approve their own team's request because approval was wrongly scoped to this
# one team too). Pulled from settings rather than hardcoded so these tests can't
# silently drift from app/auth.py.
DEPLOY_TEAM_MACHINE_GROUP_ID = get_settings().task_api_deployable_machine_group_id
# Deliberately far from DEPLOY_TEAM_MACHINE_GROUP_ID so tests can't accidentally collide
# with it regardless of what that setting's default happens to be.
REQUESTER_TEAM_MACHINE_GROUP_ID = DEPLOY_TEAM_MACHINE_GROUP_ID + 1000


@pytest.fixture()
def db_session():
    """Standalone in-memory DB for unit-testing current_deployment_status() directly,
    without going through the HTTP layer (same pattern as tests/test_reports.py)."""
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()


def _completed_request(db, *, client_id, environment, git_branch, commit_hash, requester_id, completed_at):
    request = DeploymentRequest(
        task_id="PR-1",
        client_id=client_id,
        environment=environment,
        git_branch=git_branch,
        commit_hash=commit_hash,
        requested_by=requester_id,
        status=RequestStatus.completed,
        created_at=completed_at - timedelta(hours=1),
    )
    db.add(request)
    db.flush()
    db.add(
        DeploymentExecution(
            request_id=request.id,
            executed_by=requester_id,
            claimed_at=completed_at,
            started_at=completed_at,
            completed_at=completed_at,
            status=ExecutionStatus.completed,
        )
    )
    return request


def _add_deployable_task(
    session, *, id, task_id, client_name, target, target_status="PLANNED", item_name="Some Item"
):
    task = DeployableTask(
        id=id,
        order_id=id,
        task_id=task_id,
        item_name=item_name,
        client_name=client_name,
        pos_id="0040",
        target=target,
        target_status=target_status,
    )
    session.add(task)
    return task


def _active_requests_data(response_text):
    # Pulls out list_requests()'s active_requests_json data island (request_list.html) —
    # the desktop-notification script's actual data source, independent of whichever
    # page of the table happens to be rendered. See app/routers/dashboard.py.
    match = re.search(
        r'<script type="application/json" id="active-requests-data">(.*?)</script>',
        response_text,
        re.DOTALL,
    )
    assert match, "active-requests-data script tag not found in response"
    return json.loads(match.group(1))


# --- current_deployment_status() unit tests -------------------------------------------


def test_current_deployment_status_empty_when_nothing_completed(db_session):
    assert current_deployment_status(db_session) == []


def test_current_deployment_status_returns_completed_deployment(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    _completed_request(
        db_session,
        client_id=1,
        environment=DeploymentEnvironment.live,
        git_branch="release/v12",
        commit_hash="a1b2c3d",
        requester_id=1,
        completed_at=datetime(2026, 7, 30, 9, 0, tzinfo=timezone.utc),
    )
    db_session.commit()

    rows = current_deployment_status(db_session)

    assert len(rows) == 1
    assert rows[0].client_name == "CRM"
    assert rows[0].environment == "live"
    assert rows[0].git_branch == "release/v12"
    assert rows[0].commit_hash == "a1b2c3d"
    assert rows[0].deployed_by == "Rajib Ahamad"


def test_current_deployment_status_includes_requested_by(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(User(id=1, name="Alice Requester", role=UserRole.developer))
    db_session.add(User(id=2, name="Bob Deployer", role=UserRole.developer))
    db_session.commit()
    request = DeploymentRequest(
        task_id="PR-1", client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="a1b2c3d", requested_by=1,
        status=RequestStatus.completed, created_at=datetime(2026, 7, 29, tzinfo=timezone.utc),
    )
    db_session.add(request)
    db_session.flush()
    db_session.add(
        DeploymentExecution(
            request_id=request.id,
            executed_by=2,
            claimed_at=datetime(2026, 7, 30, tzinfo=timezone.utc),
            started_at=datetime(2026, 7, 30, tzinfo=timezone.utc),
            completed_at=datetime(2026, 7, 30, tzinfo=timezone.utc),
            status=ExecutionStatus.completed,
        )
    )
    db_session.commit()

    rows = current_deployment_status(db_session)

    assert rows[0].requested_by == "Alice Requester"
    assert rows[0].deployed_by == "Bob Deployer"


def test_current_deployment_status_orders_latest_deployed_first(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(Client(id=2, name="Acme Corp"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    older = datetime(2026, 7, 1, tzinfo=timezone.utc)
    newer = datetime(2026, 7, 30, tzinfo=timezone.utc)
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v11", commit_hash="old0000", requester_id=1, completed_at=older,
    )
    _completed_request(
        db_session, client_id=2, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="new1111", requester_id=1, completed_at=newer,
    )
    db_session.commit()

    rows = current_deployment_status(db_session)

    # CRM's own deployment is older than Acme's, so Acme (the more recently deployed
    # client) sorts first, not alphabetically.
    assert [r.client_name for r in rows] == ["Acme Corp", "CRM"]


def test_current_deployment_status_picks_latest_per_client_and_environment(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    older = datetime(2026, 7, 1, tzinfo=timezone.utc)
    newer = datetime(2026, 7, 30, tzinfo=timezone.utc)
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v11", commit_hash="old0000", requester_id=1, completed_at=older,
    )
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="new1111", requester_id=1, completed_at=newer,
    )
    db_session.commit()

    rows = current_deployment_status(db_session)

    assert len(rows) == 1
    assert rows[0].git_branch == "release/v12"
    assert rows[0].commit_hash == "new1111"


def test_current_deployment_status_keeps_test_and_live_separate(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    now = datetime(2026, 7, 30, tzinfo=timezone.utc)
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.test,
        git_branch="develop", commit_hash="test0001", requester_id=1, completed_at=now,
    )
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="live0001", requester_id=1, completed_at=now,
    )
    db_session.commit()

    rows = current_deployment_status(db_session)

    assert {(r.environment, r.git_branch) for r in rows} == {("test", "develop"), ("live", "release/v12")}


def test_current_deployment_status_ignores_rejected_and_pending_requests(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    db_session.add(
        DeploymentRequest(
            task_id="PR-2",
            client_id=1,
            environment=DeploymentEnvironment.live,
            git_branch="feature/broken",
            commit_hash="dead000",
            requested_by=1,
            status=RequestStatus.rejected,
            created_at=datetime.now(timezone.utc),
        )
    )
    db_session.add(
        DeploymentRequest(
            task_id="PR-3",
            client_id=1,
            environment=DeploymentEnvironment.live,
            git_branch="feature/pending",
            commit_hash="ffff000",
            requested_by=1,
            status=RequestStatus.pending_approval,
            created_at=datetime.now(timezone.utc),
        )
    )
    db_session.commit()

    assert current_deployment_status(db_session) == []


def test_current_deployment_status_filters_by_client_environment_task_id(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(Client(id=2, name="Acme Corp"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    now = datetime(2026, 7, 30, tzinfo=timezone.utc)
    crm_live = _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="a1b2c3d", requester_id=1, completed_at=now,
    )
    crm_live.task_id = "PR-03045"
    acme_test = _completed_request(
        db_session, client_id=2, environment=DeploymentEnvironment.test,
        git_branch="develop", commit_hash="deadbee", requester_id=1, completed_at=now,
    )
    acme_test.task_id = "PR-99999"
    db_session.commit()

    assert [r.client_name for r in current_deployment_status(db_session, client_id=1)] == ["CRM"]
    assert [r.client_name for r in current_deployment_status(db_session, environment=DeploymentEnvironment.test)] == [
        "Acme Corp"
    ]
    assert [r.client_name for r in current_deployment_status(db_session, task_id="03045")] == ["CRM"]
    assert current_deployment_status(db_session, client_id=1, environment=DeploymentEnvironment.test) == []


# --- deployment_history() unit tests ---------------------------------------------------


def test_deployment_history_returns_every_deployment_not_just_latest(db_session):
    db_session.add(Client(id=1, name="CRM"))
    db_session.add(User(id=1, name="Rajib Ahamad", role=UserRole.developer))
    db_session.commit()
    older = datetime(2026, 7, 1, tzinfo=timezone.utc)
    newer = datetime(2026, 7, 30, tzinfo=timezone.utc)
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v11", commit_hash="old0000", requester_id=1, completed_at=older,
    )
    _completed_request(
        db_session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="new1111", requester_id=1, completed_at=newer,
    )
    db_session.commit()

    rows = deployment_history(db_session)

    assert [r.commit_hash for r in rows] == ["new1111", "old0000"]  # newest first, both kept


# --- Web UI (router) tests, now behind login ------------------------------------------


def test_protected_routes_redirect_anonymous_visitors_to_login(web):
    client, _session = web
    for path in ("/dashboard", "/requests", "/requests/new"):
        response = client.get(path, follow_redirects=False)
        assert response.status_code == 303, path
        assert response.headers["location"] == "/login", path


def test_home_redirects_to_dashboard_regardless_of_login(web):
    client, _session = web
    response = client.get("/", follow_redirects=False)
    assert response.status_code == 307
    assert response.headers["location"] == "/dashboard"


def test_dashboard_shows_empty_state_when_nothing_deployed(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.get("/dashboard")

    assert response.status_code == 200
    assert "No deployments have completed yet" in response.text


def _seed_two_completed_deployments(session):
    session.add(Client(id=1, name="CRM"))
    session.add(Client(id=2, name="Acme Corp"))
    session.commit()
    _completed_request(
        session, client_id=1, environment=DeploymentEnvironment.live,
        git_branch="release/v12", commit_hash="a1b2c3d", requester_id=1,
        completed_at=datetime(2026, 7, 30, tzinfo=timezone.utc),
    ).task_id = "PR-03045"
    _completed_request(
        session, client_id=2, environment=DeploymentEnvironment.test,
        git_branch="develop", commit_hash="deadbee", requester_id=1,
        completed_at=datetime(2026, 7, 29, tzinfo=timezone.utc),
    ).task_id = "PR-99999"
    session.commit()


def test_dashboard_filters_by_client(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_two_completed_deployments(session)
    login_as(client, "rajib")

    response = client.get("/dashboard", params={"client_id": "1"})

    assert response.status_code == 200
    # "Acme Corp" itself still appears in the client filter dropdown — check the row data
    # (branch/task id), not the client name, to actually confirm the table is filtered.
    assert "release/v12" in response.text
    assert "develop" not in response.text


def test_dashboard_filters_by_task_id_substring(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_two_completed_deployments(session)
    login_as(client, "rajib")

    response = client.get("/dashboard", params={"task_id": "03045"})

    assert response.status_code == 200
    assert "PR-03045" in response.text
    assert "PR-99999" not in response.text


def test_dashboard_history_lists_every_deployment(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_two_completed_deployments(session)
    login_as(client, "rajib")

    response = client.get("/dashboard/history")

    assert response.status_code == 200
    assert "CRM" in response.text
    assert "Acme Corp" in response.text


def test_dashboard_export_xlsx_returns_workbook_with_filtered_rows(web):
    from io import BytesIO

    from openpyxl import load_workbook

    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_two_completed_deployments(session)
    login_as(client, "rajib")

    response = client.get("/dashboard/export.xlsx", params={"client_id": "1"})

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "Client"  # header row
    assert rows[1][0] == "CRM"
    assert len(rows) == 2  # header + the one filtered row, Acme excluded


def test_dashboard_history_export_xlsx_includes_every_deployment(web):
    from io import BytesIO

    from openpyxl import load_workbook

    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_two_completed_deployments(session)
    login_as(client, "rajib")

    response = client.get("/dashboard/history/export.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    client_names = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}
    assert client_names == {"CRM", "Acme Corp"}


def test_new_request_form_lists_clients_and_deployable_tasks(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live")
    session.commit()
    login_as(client, "rajib")

    response = client.get("/requests/new")

    assert response.status_code == 200
    assert "CRM" in response.text
    assert "PR-03045" in response.text
    assert "+ Add new client" in response.text
    assert "Rajib Ahamad" in response.text  # static "Requested by" field, not a dropdown
    assert "requested_by" not in response.text  # no such form field anymore


def test_new_request_form_only_lists_planned_tasks(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    _add_deployable_task(session, id=100, task_id="PR-PLANNED", client_name="CRM", target="live")
    _add_deployable_task(
        session, id=101, task_id="PR-DONE", client_name="CRM", target="live", target_status="COMPLETED"
    )
    session.commit()
    login_as(client, "rajib")

    response = client.get("/requests/new")

    assert "PR-PLANNED" in response.text
    assert "PR-DONE" not in response.text


def test_new_request_form_encodes_empty_client_name_for_internal_tasks(web):
    # Regression guard for a bug where the auto-fill JS's `if (clientName)` check treated
    # an empty string the same as "no data" and skipped resetting the Client field — the
    # template must render an explicit empty data-client-name (not omit the attribute or
    # fall back to some placeholder text) so the script can tell "this task has no
    # client" apart from "no task selected yet."
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    _add_deployable_task(session, id=100, task_id="PR-INTERNAL", client_name=None, target="live")
    _add_deployable_task(session, id=101, task_id="PR-CLIENT", client_name="Acme Corp", target="test")
    session.commit()
    login_as(client, "rajib")

    response = client.get("/requests/new")

    assert 'data-id="100"' in response.text
    assert 'data-client-name=""' in response.text  # the internal task's explicit empty value
    assert 'data-id="101"' in response.text
    assert 'data-client-name="Acme Corp"' in response.text


def test_sync_deployable_tasks_now_requires_login(web):
    client, _session = web
    response = client.post("/requests/new/sync-deployable-tasks", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_sync_deployable_tasks_now_shows_success_notice(web, monkeypatch):
    # Doesn't hit the real CRM — monkeypatches the sync function itself (this
    # environment can't reach the CRM API anyway), just checks the button's redirect +
    # flash-message wiring in app/routers/dashboard.py's sync_deployable_tasks_now().
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    class FakeResult:
        total = 5

    monkeypatch.setattr("app.routers.dashboard.sync_deployable_tasks", lambda db, provider: FakeResult())

    response = client.post("/requests/new/sync-deployable-tasks", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/requests/new?synced=5"
    follow_up = client.get(response.headers["location"])
    assert "Synced 5 deployable task(s) from the CRM." in follow_up.text


def test_sync_deployable_tasks_now_shows_error_on_failure(web, monkeypatch):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    def _raise(db, provider):
        raise RuntimeError("connection refused")

    monkeypatch.setattr("app.routers.dashboard.sync_deployable_tasks", _raise)

    response = client.post("/requests/new/sync-deployable-tasks", follow_redirects=False)

    assert response.status_code == 303
    location = response.headers["location"]
    assert location.startswith("/requests/new?sync_error=")
    follow_up = client.get(location)
    assert "Could not sync from the CRM" in follow_up.text
    assert "connection refused" in follow_up.text


def test_create_request_uses_deployable_task_and_logged_in_user(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.query(DeploymentRequest).one()
    assert request.status == RequestStatus.pending_approval
    assert request.task_id == "PR-03045"  # copied from the DeployableTask, not typed
    assert request.module_name == "Some Item"  # copied from the DeployableTask's item_name
    assert request.client_id == 1
    assert request.requested_by == 1  # the logged-in user, not a form field
    assert request.version == "V12"


def test_create_request_rejects_unknown_deployable_task(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "999",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
    )

    assert response.status_code == 400
    assert "Select at least one Task ID from the list" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_request_rejects_blank_deployable_task_ids(web):
    # Mirrors what the hidden field looks like when nothing was ever successfully added
    # from the Task ID search box (request_form.html's JS).
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
    )

    assert response.status_code == 400
    assert "Select at least one Task ID from the list" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_request_combines_multiple_tasks_for_the_same_client(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live", item_name="Interface")
    _add_deployable_task(session, id=101, task_id="PR-03046", client_name="CRM", target="live", item_name="Reports")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100,101",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.query(DeploymentRequest).one()
    assert request.task_id == "PR-03045, PR-03046"
    assert request.module_name == "Interface, Reports"


def test_create_request_rejects_tasks_from_different_clients(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live")
    _add_deployable_task(session, id=101, task_id="PR-99999", client_name="Acme Corp", target="live")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100,101",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
    )

    assert response.status_code == 400
    assert "must belong to the same client" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_request_rejects_tasks_with_different_targets(web):
    # Regression test for a reported bug: the same client's Test and Live orders (e.g.
    # two DeployableTask rows both named "PR-02960 — PlanVisu" for "SCT Technology GmbH",
    # one target="test" and one target="live") passed the same-client check and got
    # combined into one request — but a request has exactly one `environment`, so this
    # would silently deploy one of the two orders to the wrong system.
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-02960", client_name="CRM", target="live")
    _add_deployable_task(session, id=101, task_id="PR-02960", client_name="CRM", target="test")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100,101",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
    )

    assert response.status_code == 400
    assert "must be for the same system" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_request_ignores_duplicate_task_ids_in_the_list(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100,100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "V12",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.query(DeploymentRequest).one()
    assert request.task_id == "PR-03045"


def test_create_request_missing_version_rerenders_form_with_error(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.add(Client(id=1, name="CRM"))
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v12",
            "commit_hash": "a1b2c3d",
            "version": "  ",
        },
    )

    assert response.status_code == 400
    assert session.query(DeploymentRequest).count() == 0


def test_create_request_with_new_client_name_creates_client_on_the_fly(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="Acme Corp", target="test")
    session.commit()
    login_as(client, "rajib")
    assert session.query(Client).count() == 0

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100",
            "client_id": "__new__",
            "new_client_name": "  Acme Corp  ",
            "environment": "test",
            "git_branch": "develop",
            "commit_hash": "abc1234",
            "version": "V1",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    new_client = session.query(Client).one()
    assert new_client.name == "Acme Corp"  # whitespace stripped
    request = session.query(DeploymentRequest).one()
    assert request.client_id == new_client.id


def test_create_request_missing_client_selection_rerenders_form_with_error(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="test")
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests",
        data={
            "deployable_task_ids": "100",
            "client_id": "",
            "environment": "test",
            "git_branch": "develop",
            "commit_hash": "abc1234",
            "version": "V1",
        },
    )

    assert response.status_code == 400
    assert "Select a client" in response.text
    assert session.query(DeploymentRequest).count() == 0


def _seed_pending_request(session):
    session.add(Client(id=1, name="CRM"))
    # Requester and Lead share a team that is deliberately NOT the deploy team — approval
    # follows the requester's own team (can_approve_deployment_request), which is a
    # different axis from who's allowed to deploy (require_deploy_team_member, scoped to
    # DEPLOY_TEAM_MACHINE_GROUP_ID specifically). Mixing these up was the actual bug.
    make_user(
        session, id=1, name="Requester", username="requester", password=DEFAULT_TEST_PASSWORD,
        machine_group_id=REQUESTER_TEAM_MACHINE_GROUP_ID,
    )
    make_user(
        session, id=2, name="Lead", role=UserRole.team_lead, username="lead", password=DEFAULT_TEST_PASSWORD,
        machine_group_id=REQUESTER_TEAM_MACHINE_GROUP_ID,
    )
    # Deployer belongs to the deploy team, not the requester's team — required for
    # deploy to succeed under require_deploy_team_member() (app/auth.py).
    make_user(
        session, id=3, name="Deployer", username="deployer", password=DEFAULT_TEST_PASSWORD,
        machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID,
    )
    session.commit()
    request = DeploymentRequest(
        task_id="PR-03045",
        module_name="Interface",
        client_id=1,
        environment=DeploymentEnvironment.live,
        git_branch="release/v12",
        commit_hash="a1b2c3d",
        version="V12",
        requested_by=1,
        status=RequestStatus.pending_approval,
        created_at=datetime.now(timezone.utc),
    )
    session.add(request)
    session.commit()
    return request


def test_requests_queue_hides_approve_reject_from_non_approvers(web):
    client, session = web
    _seed_pending_request(session)
    login_as(client, "requester")  # plain developer, not a team_lead/admin

    response = client.get("/requests")

    assert response.status_code == 200
    assert "Awaiting the requester's team lead" in response.text
    assert 'action="/requests/1/approve"' not in response.text


def test_requests_queue_shows_approve_reject_to_team_lead(web):
    client, session = web
    _seed_pending_request(session)
    login_as(client, "lead")

    response = client.get("/requests")

    assert 'action="/requests/1/approve"' in response.text
    assert 'action="/requests/1/reject"' in response.text


def test_requests_queue_active_requests_data_feeds_desktop_notifications(web):
    # The client-side JS in request_list.html reads this JSON data island (not the
    # table rows — see list_requests()'s active_requests_json, which is deliberately
    # pagination-independent) to detect two events across page reloads: a new request
    # this user can approve, and a request moving pending_approval -> approved that this
    # user can deploy.
    client, session = web
    _seed_pending_request(session)
    login_as(client, "lead")

    response = client.get("/requests")

    active = _active_requests_data(response.text)
    assert len(active) == 1
    assert active[0]["id"] == 1
    assert active[0]["status"] == "pending_approval"
    assert active[0]["canApprove"] is True
    assert active[0]["taskId"] == "PR-03045"
    assert active[0]["client"] == "CRM"


def test_requests_queue_shows_module_name_and_version_columns(web):
    client, session = web
    _seed_pending_request(session)
    login_as(client, "lead")

    response = client.get("/requests")

    assert "<th>Module Name</th>" in response.text
    assert "<th>Version</th>" in response.text
    assert "<td>Interface</td>" in response.text
    assert "<td>V12</td>" in response.text


def test_requests_queue_branch_commit_has_view_button_for_full_text(web):
    # A long branch name used to force the whole table into horizontal scroll — the cell
    # is now truncated with ellipsis (branch-commit-preview, style.css) and this button
    # reveals the untruncated text in the shared modal (base.html), same pattern as the
    # Changes column but with its own title.
    client, session = web
    _seed_pending_request(session)  # git_branch="release/v12", commit_hash="a1b2c3d"
    login_as(client, "lead")

    response = client.get("/requests")

    assert 'class="branch-commit-cell"' in response.text
    assert 'class="link-button view-detail"' in response.text
    assert 'data-detail-title="Branch Name / Commit"' in response.text
    assert 'data-detail="release/v12 / a1b2c3d"' in response.text


def _seed_many_requests(session, count):
    for i in range(count):
        session.add(
            DeploymentRequest(
                task_id=f"PR-{i:03d}",
                requested_by=1,
                status=RequestStatus.rejected,
                # Descending created_at as i increases, so i=0 is newest (sorts first —
                # list_requests() orders by created_at desc) and i=count-1 is oldest.
                created_at=datetime.now(timezone.utc) - timedelta(minutes=i),
            )
        )
    session.commit()


def test_requests_queue_paginates_at_15_per_page_by_default(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_many_requests(session, 20)  # 15 on page 1, 5 on page 2 — unambiguous at the default
    login_as(client, "rajib")

    page1 = client.get("/requests")
    assert page1.status_code == 200
    assert "Page 1 of 2 (20 total)" in page1.text
    assert "PR-000" in page1.text  # newest — first page
    assert "PR-014" in page1.text  # 15th item — still first page
    assert "PR-019" not in page1.text  # oldest — second page

    page2 = client.get("/requests", params={"page": 2})
    assert page2.status_code == 200
    assert "Page 2 of 2 (20 total)" in page2.text
    assert "PR-019" in page2.text
    assert "PR-000" not in page2.text


def test_requests_queue_page_size_is_configurable(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_many_requests(session, 20)
    login_as(client, "rajib")

    response = client.get("/requests", params={"page_size": 25})

    assert response.status_code == 200
    # Everything fits on one page at 25/page, so the pagination nav itself doesn't render
    # — but every row, including the one that was pushed to page 2 at the 15/page default,
    # must show up.
    assert 'class="pagination"' not in response.text
    assert "PR-000" in response.text
    assert "PR-019" in response.text


def test_requests_queue_rejects_out_of_range_page_size(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_many_requests(session, 20)
    login_as(client, "rajib")

    # Not one of the dropdown's allowed values (15/25/50/100) — falls back to the
    # default rather than letting a hand-edited URL request an arbitrarily large page.
    response = client.get("/requests", params={"page_size": 999})

    assert response.status_code == 200
    assert "Page 1 of 2 (20 total)" in response.text  # confirms it fell back to 15/page


def test_requester_can_delete_own_pending_request(web):
    client, session = web
    _seed_pending_request(session)  # request id=1, requested_by=1 ("requester")
    login_as(client, "requester")

    response = client.post("/requests/1/delete", follow_redirects=False)

    assert response.status_code == 303
    assert session.get(DeploymentRequest, 1) is None


def test_admin_can_delete_any_pending_request(web):
    client, session = web
    _seed_pending_request(session)
    make_user(session, id=5, name="Root Admin", role=UserRole.admin, username="root", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "root")

    response = client.post("/requests/1/delete", follow_redirects=False)

    assert response.status_code == 303
    assert session.get(DeploymentRequest, 1) is None


def test_other_user_cannot_delete_someone_elses_request(web):
    client, session = web
    _seed_pending_request(session)
    login_as(client, "lead")  # team_lead, but not the requester and not an admin

    response = client.post("/requests/1/delete")

    assert response.status_code == 403
    assert session.get(DeploymentRequest, 1) is not None


def test_cannot_delete_a_request_once_deployment_has_started(web):
    # DELETABLE_REQUEST_STATUSES (app/models/deployment_request.py) deliberately excludes
    # in_progress/completed — deleting either would silently erase real execution history.
    client, session = web
    _seed_pending_request(session)
    request = session.get(DeploymentRequest, 1)
    request.status = RequestStatus.in_progress
    session.commit()
    login_as(client, "requester")

    response = client.post("/requests/1/delete")

    assert response.status_code == 403
    assert session.get(DeploymentRequest, 1) is not None


def test_delete_request_requires_login(web):
    client, session = web
    _seed_pending_request(session)

    response = client.post("/requests/1/delete", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/login"
    assert session.get(DeploymentRequest, 1) is not None


def _seed_editable_request(session):
    """Like _seed_pending_request, but also backs the request with a real DeployableTask
    (id=100) referenced via deployable_task_ids, so its edit form's Task ID picker has
    something real to prefill/re-validate against."""
    request = _seed_pending_request(session)
    _add_deployable_task(session, id=100, task_id="PR-03045", client_name="CRM", target="live")
    request.deployable_task_ids = "100"
    session.commit()
    return request


def test_requester_can_view_edit_form_for_own_pending_request(web):
    client, session = web
    _seed_editable_request(session)
    login_as(client, "requester")

    response = client.get("/requests/1/edit")

    assert response.status_code == 200
    assert 'value="release/v12"' in response.text


def test_other_user_cannot_view_edit_form(web):
    client, session = web
    _seed_editable_request(session)
    login_as(client, "lead")  # team_lead, but not the requester and not an admin

    response = client.get("/requests/1/edit")

    assert response.status_code == 403


def test_cannot_view_edit_form_once_approved(web):
    client, session = web
    request = _seed_editable_request(session)
    request.status = RequestStatus.approved
    session.commit()
    login_as(client, "requester")

    response = client.get("/requests/1/edit")

    assert response.status_code == 403


def test_requester_can_edit_own_pending_request(web):
    client, session = web
    _seed_editable_request(session)
    login_as(client, "requester")

    response = client.post(
        "/requests/1/edit",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v13",
            "commit_hash": "e5f6g7h",
            "version": "V13",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.get(DeploymentRequest, 1)
    assert request.git_branch == "release/v13"
    assert request.commit_hash == "e5f6g7h"
    assert request.version == "V13"
    assert request.status == RequestStatus.pending_approval  # editing doesn't change status


def test_admin_can_edit_any_pending_request(web):
    client, session = web
    _seed_editable_request(session)
    make_user(session, id=5, name="Root Admin", role=UserRole.admin, username="root", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "root")

    response = client.post(
        "/requests/1/edit",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v13",
            "commit_hash": "e5f6g7h",
            "version": "V13",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert session.get(DeploymentRequest, 1).git_branch == "release/v13"


def test_other_user_cannot_edit_someone_elses_request(web):
    client, session = web
    _seed_editable_request(session)
    login_as(client, "lead")  # team_lead, but not the requester and not an admin

    response = client.post(
        "/requests/1/edit",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v13",
            "commit_hash": "e5f6g7h",
            "version": "V13",
        },
    )

    assert response.status_code == 403
    assert session.get(DeploymentRequest, 1).git_branch == "release/v12"


def test_cannot_edit_a_request_once_approved(web):
    client, session = web
    request = _seed_editable_request(session)
    request.status = RequestStatus.approved
    session.commit()
    login_as(client, "requester")

    response = client.post(
        "/requests/1/edit",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v13",
            "commit_hash": "e5f6g7h",
            "version": "V13",
        },
    )

    assert response.status_code == 403
    assert session.get(DeploymentRequest, 1).git_branch == "release/v12"


def test_cannot_edit_a_rejected_request(web):
    # Confirmed with the user: rejected is a decided outcome, same as approved — a
    # rejected request needs a brand new submission, not an edit-and-resubmit.
    client, session = web
    request = _seed_editable_request(session)
    request.status = RequestStatus.rejected
    session.commit()
    login_as(client, "requester")

    response = client.post(
        "/requests/1/edit",
        data={
            "deployable_task_ids": "100",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v13",
            "commit_hash": "e5f6g7h",
            "version": "V13",
        },
    )

    assert response.status_code == 403


def test_edit_request_rejects_tasks_with_different_targets(web):
    # Same server-side backstop create_request() applies (test_create_request_rejects_
    # tasks_with_different_targets above) must still hold on edit.
    client, session = web
    request = _seed_editable_request(session)
    _add_deployable_task(session, id=101, task_id="PR-02960", client_name="CRM", target="test")
    session.commit()
    login_as(client, "requester")

    response = client.post(
        "/requests/1/edit",
        data={
            "deployable_task_ids": "100,101",
            "client_id": "1",
            "environment": "live",
            "git_branch": "release/v13",
            "commit_hash": "e5f6g7h",
            "version": "V13",
        },
    )

    assert response.status_code == 400
    assert "must be for the same system" in response.text
    assert session.get(DeploymentRequest, request.id).git_branch == "release/v12"


def test_db_dump_restore_request_cannot_be_edited(web):
    # Editing only exists for `standard` requests — db_dump_restore/test_local requests
    # are created straight into `approved` and never have a real editable window, but
    # this pins the request_type restriction independent of status too.
    client, session = web
    make_user(session, id=1, name="Requester", username="requester", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    request = DeploymentRequest(
        request_type=RequestType.db_dump_restore,
        dump_source="crm-live DB",
        version="V12",
        requested_by=1,
        status=RequestStatus.pending_approval,  # artificial, to isolate the type check
        created_at=datetime.now(timezone.utc),
    )
    session.add(request)
    session.commit()
    login_as(client, "requester")

    response = client.get(f"/requests/{request.id}/edit")

    assert response.status_code == 403


def test_edit_request_requires_login(web):
    client, session = web
    _seed_editable_request(session)

    response = client.get("/requests/1/edit", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_requests_queue_shows_delete_button_only_to_requester_and_admin(web):
    client, session = web
    _seed_pending_request(session)

    login_as(client, "requester")
    as_requester = client.get("/requests")
    assert 'action="/requests/1/delete"' in as_requester.text

    login_as(client, "lead")
    as_other_user = client.get("/requests")
    assert 'action="/requests/1/delete"' not in as_other_user.text


def test_requests_queue_clamps_out_of_range_page(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_many_requests(session, 3)
    login_as(client, "rajib")

    response = client.get("/requests", params={"page": 999})

    # Only one page exists, so the nav itself doesn't render (same as any single-page
    # queue) — but the request must not 500/422, and page clamps down to showing
    # everything rather than an empty "page 999 of 1" table.
    assert response.status_code == 200
    assert 'class="pagination"' not in response.text
    assert "PR-000" in response.text


def test_requests_queue_hides_pagination_nav_when_everything_fits_on_one_page(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    _seed_many_requests(session, 3)
    login_as(client, "rajib")

    response = client.get("/requests")

    assert response.status_code == 200
    assert 'class="pagination"' not in response.text


def test_requests_queue_hides_approve_reject_from_team_lead_of_a_different_team(web):
    client, session = web
    _seed_pending_request(session)
    make_user(
        session, id=4, name="Other Team Lead", role=UserRole.team_lead, username="otherlead",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=REQUESTER_TEAM_MACHINE_GROUP_ID + 1,
    )
    session.commit()
    login_as(client, "otherlead")

    response = client.get("/requests")

    assert "Awaiting the requester's team lead" in response.text
    assert 'action="/requests/1/approve"' not in response.text


def test_approve_requires_approver_role(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "requester")  # plain developer

    response = client.post(f"/requests/{request.id}/approve")

    assert response.status_code == 403
    session.refresh(request)
    assert request.status == RequestStatus.pending_approval


def test_approve_rejects_team_lead_of_a_different_team(web):
    # Only the REQUESTER's own team lead may approve — not just any team_lead in the CRM
    # roster, and specifically not a team_lead of some unrelated team (including the
    # deploy team itself) — app/auth.py's can_approve_deployment_request().
    client, session = web
    request = _seed_pending_request(session)
    make_user(
        session, id=4, name="Other Team Lead", role=UserRole.team_lead, username="otherlead",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=REQUESTER_TEAM_MACHINE_GROUP_ID + 1,
    )
    session.commit()
    login_as(client, "otherlead")

    response = client.post(f"/requests/{request.id}/approve")

    assert response.status_code == 403
    session.refresh(request)
    assert request.status == RequestStatus.pending_approval


def test_approve_rejects_deploy_team_lead_who_is_not_the_requesters_own_lead(web):
    # Regression test for the actual reported bug: being a team_lead who belongs to the
    # deploy team does NOT make you the requester's team lead — approval must follow the
    # requester's own team, not the deploy team (app/auth.py's
    # can_approve_deployment_request(), which is deliberately independent of
    # require_deploy_team_member()).
    client, session = web
    request = _seed_pending_request(session)
    make_user(
        session, id=8, name="Deploy Team Lead", role=UserRole.team_lead, username="deployteamlead",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID,
    )
    session.commit()
    login_as(client, "deployteamlead")

    response = client.post(f"/requests/{request.id}/approve")

    assert response.status_code == 403
    session.refresh(request)
    assert request.status == RequestStatus.pending_approval


def test_approve_allows_admin_regardless_of_team(web):
    client, session = web
    request = _seed_pending_request(session)
    make_user(
        session, id=5, name="Root Admin", role=UserRole.admin, username="root",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=None,
    )
    session.commit()
    login_as(client, "root")

    response = client.post(f"/requests/{request.id}/approve", follow_redirects=False)

    assert response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.approved


def test_start_requires_deploy_team_membership(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")
    make_user(
        session, id=6, name="Outside Developer", username="outsider",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID + 1,
    )
    session.commit()
    login_as(client, "outsider")

    response = client.post(f"/requests/{request.id}/start")

    assert response.status_code == 403
    session.refresh(request)
    assert request.status == RequestStatus.approved


def test_start_before_approval_is_rejected_with_409(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "deployer")

    response = client.post(f"/requests/{request.id}/start")

    assert response.status_code == 409
    session.refresh(request)
    assert request.status == RequestStatus.pending_approval


def test_start_moves_request_to_in_progress_and_records_who(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")
    login_as(client, "deployer")

    response = client.post(f"/requests/{request.id}/start", follow_redirects=False)

    assert response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.in_progress

    execution = session.query(DeploymentExecution).filter_by(request_id=request.id).one()
    assert execution.executed_by == 3  # "deployer"
    assert execution.status == ExecutionStatus.in_progress
    assert execution.started_at is not None
    assert execution.completed_at is None


def _seed_in_progress_standard_request(session, *, client_id=1, environment=DeploymentEnvironment.live):
    session.add(Client(id=client_id, name="CRM"))
    make_user(session, id=1, name="Requester", username="requester", password=DEFAULT_TEST_PASSWORD)
    make_user(
        session, id=3, name="Deployer", username="deployer", password=DEFAULT_TEST_PASSWORD,
        machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID,
    )
    session.commit()
    request = DeploymentRequest(
        id=1, request_type=RequestType.standard, client_id=client_id, environment=environment,
        git_branch="release/v12", commit_hash="a1b2c3d", version="V12", requested_by=1,
        status=RequestStatus.in_progress, created_at=datetime.now(timezone.utc),
    )
    session.add(request)
    session.add(
        DeploymentExecution(
            request_id=1, executed_by=3, claimed_at=datetime.now(timezone.utc),
            started_at=datetime.now(timezone.utc), status=ExecutionStatus.in_progress,
        )
    )
    session.commit()
    return request


def test_deploy_request_requires_current_version_for_standard_requests(web):
    client, session = web
    _seed_in_progress_standard_request(session)
    login_as(client, "deployer")

    response = client.post("/requests/1/deploy", data={"current_version": ""})

    assert response.status_code == 400
    assert session.get(DeploymentRequest, 1).status == RequestStatus.in_progress
    assert session.query(ClientVersionRecord).count() == 0


def test_deploy_request_creates_client_version_record(web):
    client, session = web
    _seed_in_progress_standard_request(session)
    session.add(BitbucketMainBranchStatus(id=1, version="2026.34.40", pr_number=1234))
    session.commit()
    login_as(client, "deployer")

    response = client.post("/requests/1/deploy", data={"current_version": "2026.34.34"}, follow_redirects=False)

    assert response.status_code == 303
    assert session.get(DeploymentRequest, 1).status == RequestStatus.completed
    record = session.query(ClientVersionRecord).one()
    assert record.client_id == 1
    assert record.environment == DeploymentEnvironment.live
    assert record.current_version == "2026.34.34"
    assert record.previous_version is None
    assert record.main_version == "2026.34.40"
    assert record.main_pr_number == 1234
    assert record.deployment_request_id == 1
    assert record.recorded_by == 3


def test_deploy_request_fills_previous_version_from_prior_record(web):
    client, session = web
    _seed_in_progress_standard_request(session)
    session.add(
        ClientVersionRecord(
            client_id=1, environment=DeploymentEnvironment.live, current_version="2026.34.30",
            deployment_request_id=1, recorded_by=3,
            created_at=datetime.now(timezone.utc), updated_at=datetime.now(timezone.utc),
        )
    )
    session.commit()
    login_as(client, "deployer")

    client.post("/requests/1/deploy", data={"current_version": "2026.34.34"})

    latest = session.query(ClientVersionRecord).order_by(ClientVersionRecord.id.desc()).first()
    assert latest.previous_version == "2026.34.30"
    assert latest.current_version == "2026.34.34"


def test_deploy_request_works_without_a_bitbucket_sync_yet(web):
    # No BitbucketMainBranchStatus row at all — main_version/main_pr_number just
    # come through null rather than erroring.
    client, session = web
    _seed_in_progress_standard_request(session)
    login_as(client, "deployer")

    response = client.post("/requests/1/deploy", data={"current_version": "2026.34.34"}, follow_redirects=False)

    assert response.status_code == 303
    record = session.query(ClientVersionRecord).one()
    assert record.main_version is None
    assert record.main_pr_number is None


def test_requests_queue_renders_deploy_version_dialog_for_standard_in_progress_row(web):
    client, session = web
    _seed_in_progress_standard_request(session)
    session.add(
        ClientVersionRecord(
            client_id=1, environment=DeploymentEnvironment.live, current_version="2026.34.30",
            deployment_request_id=1, recorded_by=3,
            created_at=datetime.now(timezone.utc), updated_at=datetime.now(timezone.utc),
        )
    )
    session.commit()
    login_as(client, "deployer")

    response = client.get("/requests")

    assert response.status_code == 200
    assert 'id="deploy-version-modal"' in response.text
    assert 'data-deploy-request-id="1"' in response.text
    assert 'data-deploy-previous-version="2026.34.30"' in response.text
    # The button itself is no longer a bare submit for standard rows:
    assert 'type="button" class="deploy" data-deploy-request-id="1"' in response.text


def test_deploy_before_start_is_rejected_with_409(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")
    login_as(client, "deployer")

    response = client.post(f"/requests/{request.id}/deploy")

    assert response.status_code == 409
    session.refresh(request)
    assert request.status == RequestStatus.approved


def test_deploy_requires_deploy_team_membership(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")
    login_as(client, "deployer")
    client.post(f"/requests/{request.id}/start")
    make_user(
        session, id=6, name="Outside Developer", username="outsider",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID + 1,
    )
    session.commit()
    login_as(client, "outsider")

    response = client.post(f"/requests/{request.id}/deploy")

    assert response.status_code == 403
    session.refresh(request)
    assert request.status == RequestStatus.in_progress


def test_deploy_allows_admin_regardless_of_team(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")
    make_user(
        session, id=7, name="Root Admin", role=UserRole.admin, username="root2",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=None,
    )
    session.commit()
    login_as(client, "root2")
    client.post(f"/requests/{request.id}/start")

    response = client.post(
        f"/requests/{request.id}/deploy", data={"current_version": "2026.34.34"}, follow_redirects=False
    )

    assert response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.completed


def test_deploy_can_be_done_by_a_different_deploy_team_member_than_who_started_it(web):
    # "Any member of the deploy team" (project_plan.md Section 3), not a personal claim
    # lock — starting and completing don't have to be the same person.
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")
    login_as(client, "deployer")
    client.post(f"/requests/{request.id}/start")
    make_user(
        session, id=9, name="Second Deployer", username="deployer2",
        password=DEFAULT_TEST_PASSWORD, machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID,
    )
    session.commit()
    login_as(client, "deployer2")

    response = client.post(
        f"/requests/{request.id}/deploy", data={"current_version": "2026.34.34"}, follow_redirects=False
    )

    assert response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.completed
    execution = session.query(DeploymentExecution).filter_by(request_id=request.id).one()
    assert execution.executed_by == 3  # still attributed to whoever started it
    assert execution.completed_at is not None


def test_approve_then_start_then_deploy_updates_dashboard_with_logged_in_users(web):
    client, session = web
    request = _seed_pending_request(session)

    login_as(client, "lead")
    approve_response = client.post(f"/requests/{request.id}/approve", follow_redirects=False)
    assert approve_response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.approved

    login_as(client, "deployer")
    start_response = client.post(f"/requests/{request.id}/start", follow_redirects=False)
    assert start_response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.in_progress
    assert "In Progress" in client.get("/requests").text

    deploy_response = client.post(
        f"/requests/{request.id}/deploy", data={"current_version": "2026.34.34"}, follow_redirects=False
    )
    assert deploy_response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.completed

    execution = session.query(DeploymentExecution).filter_by(request_id=request.id).one()
    assert execution.executed_by == 3  # "deployer", not a form-picked name

    dashboard_response = client.get("/dashboard")
    assert "CRM" in dashboard_response.text
    assert "release/v12" in dashboard_response.text
    assert "a1b2c3d" in dashboard_response.text
    assert "Deployer" in dashboard_response.text


def test_reject_marks_request_rejected_and_keeps_it_off_the_dashboard(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")

    response = client.post(
        f"/requests/{request.id}/reject", data={"comment": "wrong branch"}, follow_redirects=False
    )

    assert response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.rejected

    dashboard_response = client.get("/dashboard")
    assert "No deployments have completed yet" in dashboard_response.text


def test_approve_twice_is_rejected_with_409(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "lead")
    client.post(f"/requests/{request.id}/approve")

    second_attempt = client.post(f"/requests/{request.id}/approve")

    assert second_attempt.status_code == 409


def test_deploy_before_approval_is_rejected_with_409(web):
    client, session = web
    request = _seed_pending_request(session)
    login_as(client, "deployer")

    response = client.post(f"/requests/{request.id}/deploy")

    assert response.status_code == 409


def test_approve_unknown_request_returns_404(web):
    client, session = web
    make_user(
        session, id=2, name="Lead", role=UserRole.team_lead, username="lead", password=DEFAULT_TEST_PASSWORD,
        machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID,
    )
    session.commit()
    login_as(client, "lead")

    response = client.post("/requests/999/approve")

    assert response.status_code == 404


# --- Database Dump & Restore requests (no approval required) -------------------------


def test_new_request_form_shows_test_local_server_suggestions(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.get("/requests/new")

    assert response.status_code == 200
    assert "crm.test.local" in response.text
    assert "tmp.test.local" in response.text
    assert "vop.test.local" in response.text


def test_create_db_dump_restore_request_with_restore_source_skips_approval(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/db-dump-restore",
        data={"dump_source": "crm-live", "version": "V12", "restore_source": "crm-staging"},
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.query(DeploymentRequest).one()
    assert request.request_type == RequestType.db_dump_restore
    assert request.dump_source == "crm-live"
    assert request.version == "V12"
    assert request.restore_source == "crm-staging"
    assert request.share_with_requestor is False
    assert request.requested_by == 1
    # No approval required — lands straight in the deploy team's queue.
    assert request.status == RequestStatus.approved


def test_requests_queue_row_for_db_dump_restore_carries_deploy_notification_data(web):
    # Regression test: db_dump_restore (and test_local) requests skip the approval stage
    # and land straight in `approved` — the notification JS's "new pending_approval"
    # trigger never sees them arrive, and the "pending_approval -> approved" transition
    # trigger never fires either, since they never pass through pending_approval at all.
    # The only way the deploy team hears about a brand new one is a THIRD trigger keyed
    # off requestType/dumpSource in active_requests_json, which this asserts are present.
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")
    client.post(
        "/requests/db-dump-restore",
        data={"dump_source": "crm-live", "version": "V12", "restore_source": "crm-staging"},
    )

    response = client.get("/requests")

    active = _active_requests_data(response.text)
    assert len(active) == 1
    assert active[0]["status"] == "approved"
    assert active[0]["requestType"] == "db_dump_restore"
    assert active[0]["dumpSource"] == "crm-live"
    # Not sourced from deployable_tasks, so there's no module name to show — but the
    # request's own `version` (which used to be squeezed into the Details column as a
    # badge) still shows in the dedicated Version column.
    assert "<td>—</td>" in response.text
    assert "<td>V12</td>" in response.text


def test_create_db_dump_restore_request_with_share_with_requestor(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/db-dump-restore",
        data={"dump_source": "crm-live", "version": "V10", "share_with_requestor": "on"},
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.query(DeploymentRequest).one()
    assert request.share_with_requestor is True
    assert request.restore_source is None
    assert request.version == "V10"
    assert request.status == RequestStatus.approved


def test_create_db_dump_restore_request_requires_dump_source(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/db-dump-restore",
        data={"dump_source": "  ", "version": "V12", "restore_source": "crm-staging"},
    )

    assert response.status_code == 400
    assert "Dump source is required" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_db_dump_restore_request_requires_version(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/db-dump-restore",
        data={"dump_source": "crm-live", "version": "  ", "restore_source": "crm-staging"},
    )

    assert response.status_code == 400
    assert "Application version is required" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_db_dump_restore_request_rejects_both_restore_and_share(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/db-dump-restore",
        data={
            "dump_source": "crm-live",
            "version": "V12",
            "restore_source": "crm-staging",
            "share_with_requestor": "on",
        },
    )

    assert response.status_code == 400
    assert "not both" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_db_dump_restore_request_rejects_neither_restore_nor_share(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/db-dump-restore", data={"dump_source": "crm-live", "version": "V12"}
    )

    assert response.status_code == 400
    assert "Provide a restore source" in response.text
    assert session.query(DeploymentRequest).count() == 0


# --- Test.local deployment requests (no approval required) ---------------------------


def test_create_test_local_request_skips_approval(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/test-local",
        data={
            "server": "crm.test.local",
            "git_branch": "feature/my-branch",
            "version": "V12",
            "changes_description": "quick check",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    request = session.query(DeploymentRequest).one()
    assert request.request_type == RequestType.test_local
    assert request.server == "crm.test.local"
    assert request.git_branch == "feature/my-branch"
    assert request.version == "V12"
    assert request.changes_description == "quick check"
    assert request.requested_by == 1
    assert request.status == RequestStatus.approved


def test_create_test_local_request_rejects_non_test_local_host(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/test-local",
        data={"server": "crm-live.example.com", "git_branch": "develop", "version": "V1"},
    )

    assert response.status_code == 400
    assert "test.local" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_test_local_request_requires_branch(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/test-local", data={"server": "crm.test.local", "git_branch": "  ", "version": "V1"}
    )

    assert response.status_code == 400
    assert "Branch name is required" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_create_test_local_request_requires_version(web):
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    session.commit()
    login_as(client, "rajib")

    response = client.post(
        "/requests/test-local", data={"server": "crm.test.local", "git_branch": "develop", "version": "  "}
    )

    assert response.status_code == 400
    assert "Application version is required" in response.text
    assert session.query(DeploymentRequest).count() == 0


def test_no_approval_required_requests_go_straight_to_deploy_queue(web):
    # The whole point of these two request types: they never touch pending_approval, so
    # a deploy-team member sees "Start Deployment" immediately, with no team-lead step at
    # all — then the usual start -> deploy sequence applies from there.
    client, session = web
    make_user(session, id=1, name="Rajib Ahamad", username="rajib", password=DEFAULT_TEST_PASSWORD)
    make_user(
        session, id=2, name="Deployer", username="deployer", password=DEFAULT_TEST_PASSWORD,
        machine_group_id=DEPLOY_TEAM_MACHINE_GROUP_ID,
    )
    session.commit()
    login_as(client, "rajib")
    client.post("/requests/test-local", data={"server": "tmp.test.local", "git_branch": "develop", "version": "V1"})

    login_as(client, "deployer")
    response = client.get("/requests")

    assert response.status_code == 200
    assert "Awaiting the requester's team lead" not in response.text
    request = session.query(DeploymentRequest).one()
    assert f'action="/requests/{request.id}/start"' in response.text

    start_response = client.post(f"/requests/{request.id}/start", follow_redirects=False)
    assert start_response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.in_progress

    deploy_response = client.post(f"/requests/{request.id}/deploy", follow_redirects=False)
    assert deploy_response.status_code == 303
    session.refresh(request)
    assert request.status == RequestStatus.completed
